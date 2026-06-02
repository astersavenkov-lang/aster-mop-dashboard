"""
ADM-ASTER · Telegram Bot · Авторассылка отчётов по брендам
===========================================================
Бот:
  1. Мониторит группу-источник
  2. При получении Excel файла — анализирует данные
  3. Генерирует HTML отчёт через Claude API
  4. Отправляет HTML каждому бренду в свою группу

Установка (один раз):
  pip install python-telegram-bot anthropic openpyxl

Запуск:
  python aster_bot.py
"""

import logging
import io
import json
from datetime import datetime
from pathlib import Path
from collections import defaultdict

import anthropic
from openpyxl import load_workbook
from telegram import Update, Document
from telegram.ext import (
    Application, MessageHandler, filters, ContextTypes
)
from telegram.constants import ParseMode

# ══════════════════════════════════════════════════════════════
#  НАСТРОЙКИ — заполните перед запуском
# ══════════════════════════════════════════════════════════════

BOT_TOKEN = "ВАШ_ТОКЕН_ОТ_BOTFATHER"   # Получить у @BotFather

ANTHROPIC_KEY = "sk-ant-..."            # console.anthropic.com → API Keys

# ID группы-источника (откуда приходит Excel)
# Как узнать ID: добавьте бота в группу, перешлите любое сообщение
# боту @userinfobot — он покажет ID группы (отрицательное число)
SOURCE_GROUP_ID = -1001234567890        # ← замените на ID вашей группы

# Группы-получатели по брендам
# Формат: "Название бренда": ID_группы
# Бот должен быть добавлен в каждую группу с правом отправки сообщений
BRAND_GROUPS = {
    "Chery":      -1001111111111,   # ← замените на реальные ID групп
    "Chevrolet":  -1001222222222,
    "Haval":      -1001333333333,
    "KIA":        -1001444444444,
    "BYD":        -1001555555555,
    "Lada":       -1001666666666,
    "Jetour":     -1001777777777,
    "Changan":    -1001888888888,
    "Другие":     -1001999999999,   # Все остальные бренды
}

# Группа для логов/уведомлений об ошибках (можно свой личный чат)
# Получить свой ID: напишите @userinfobot
ADMIN_CHAT_ID = 123456789           # ← ваш личный Telegram ID

# ══════════════════════════════════════════════════════════════

logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(message)s",
    level=logging.INFO,
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("aster_bot.log", encoding="utf-8"),
    ]
)
log = logging.getLogger(__name__)


# ─── Определение бренда по названию филиала ──────────────────

def get_brand(branch_name: str) -> str:
    n = str(branch_name).upper()
    if "CHEVROLET" in n: return "Chevrolet"
    if "BYD"       in n: return "BYD"
    if "KIA"       in n: return "KIA"
    if "CHERY"     in n: return "Chery"
    if "HAVAL"     in n: return "Haval"
    if "JETOUR"    in n: return "Jetour"
    if "CHANGAN"   in n: return "Changan"
    if "LADA"      in n: return "Lada"
    if "HYUNDAI"   in n: return "Hyundai"
    if "DEEPAL"    in n: return "Другие"
    if "DONGFENG"  in n: return "Другие"
    if "LI "       in n: return "Другие"
    return "Другие"


# ─── Чтение Excel ────────────────────────────────────────────

def parse_excel(file_bytes: bytes) -> dict:
    """
    Читает Excel и возвращает dict:
    {
      "all_rows": [...],           # все строки
      "by_brand": {                # сгруппировано по брендам
        "Chery": [...],
        "Haval": [...],
        ...
      },
      "totals": {...}              # итоговая строка
    }
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    # Заголовки
    headers = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(rows[0])]

    data_rows = []
    totals_row = None

    for row in rows[1:]:
        if not any(row):
            continue
        d = {}
        for i, val in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            # Чистим float экспоненты типа 0E-20
            if isinstance(val, float) and abs(val) < 1e-10:
                val = 0.0
            d[key] = val

        # Определяем строку итогов
        first_val = str(row[0] or "").lower()
        if "итого" in first_val or "сумма" in first_val or "total" in first_val:
            totals_row = d
        elif any(v for v in row if v is not None and v != ""):
            data_rows.append(d)

    # Группируем по брендам
    by_brand = defaultdict(list)
    for row in data_rows:
        # Пробуем найти колонку с филиалом
        branch = (
            row.get("Филиал") or
            row.get("МОП") and "" or
            list(row.values())[1] if len(row) > 1 else ""
        )
        brand = get_brand(str(branch or ""))
        by_brand[brand].append(row)

    wb.close()
    return {
        "headers": headers,
        "all_rows": data_rows,
        "by_brand": dict(by_brand),
        "totals": totals_row,
    }


# ─── Генерация HTML через Claude ─────────────────────────────

SYSTEM_PROMPT = """Ты — Senior Business Analyst ADM-ASTER.
Создаёшь компактные HTML отчёты для менеджеров дилерской сети.
Стиль: тёмная шапка (#0D1B2A), золотой акцент (#E8A838), navy (#1B3A5C).
Всегда возвращай ТОЛЬКО валидный HTML без markdown-обёртки и без ```."""

def build_prompt_brand(brand: str, rows: list, totals: dict, report_date: str) -> str:
    rows_json = json.dumps(rows[:120], ensure_ascii=False)
    totals_str = json.dumps(totals or {}, ensure_ascii=False)

    return f"""Создай HTML отчёт по бренду {brand} для ADM-ASTER.

Дата: {report_date}
Бренд: {brand}
Данные строк: {rows_json}
Итого по сети: {totals_str}

ТРЕБОВАНИЯ:
1. Шапка: логотип | ASTER, название бренда, дата, период
2. KPI-плашки вверху: РЛ всего, Визиты, Тест-драйвы, Продажи, % конверсии, % упущенных
3. Таблица по филиалам с цветовой индикацией:
   - % конверсии: ≥3% зелёный, 1-3% оранжевый, <1% красный
   - % упущенных: <10% зелёный, 10-30% оранжевый, >30% красный
4. Воронка (прогресс-бары): РЛ → Визиты → ТД → Продажи
5. Раздел «Выводы»: топ-2 лучших и топ-2 проблемных филиала
6. Рекомендации: 3-4 конкретных действия
7. Компактный дизайн, мобильный (max-width 700px, font 12-13px)
8. Цвета ADM-ASTER: bg #0D1B2A, navy #1B3A5C, accent #2E6DA4, gold #E8A838

Верни ТОЛЬКО HTML, без markdown."""


def generate_brand_html(brand: str, rows: list, totals: dict) -> str:
    client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
    report_date = datetime.now().strftime("%d.%m.%Y %H:%M")

    prompt = build_prompt_brand(brand, rows, totals, report_date)

    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=6000,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": prompt}],
    )
    return message.content[0].text


# ─── Обработчик входящих файлов ──────────────────────────────

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.effective_message
    chat_id = update.effective_chat.id

    # Принимаем файлы ТОЛЬКО из группы-источника
    if chat_id != SOURCE_GROUP_ID:
        return

    doc: Document = msg.document
    if not doc:
        return

    filename = doc.file_name or ""
    # Проверяем что это Excel
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        return

    log.info(f"📥 Получен файл: {filename} из группы {chat_id}")

    # Уведомляем администратора
    await context.bot.send_message(
        ADMIN_CHAT_ID,
        f"📥 Получен файл: `{filename}`\nНачинаю обработку...",
        parse_mode=ParseMode.MARKDOWN,
    )

    try:
        # Скачиваем файл
        tg_file = await doc.get_file()
        file_bytes = await tg_file.download_as_bytearray()
        log.info(f"  Размер файла: {len(file_bytes):,} байт")

        # Парсим Excel
        await context.bot.send_message(ADMIN_CHAT_ID, "📊 Читаю Excel...")
        parsed = parse_excel(bytes(file_bytes))

        if not parsed or not parsed.get("all_rows"):
            await context.bot.send_message(ADMIN_CHAT_ID, "❌ Файл пустой или неверный формат")
            return

        brands_found = list(parsed["by_brand"].keys())
        total_rows = len(parsed["all_rows"])
        log.info(f"  Строк: {total_rows}, Брендов: {brands_found}")
        await context.bot.send_message(
            ADMIN_CHAT_ID,
            f"✅ Прочитано: {total_rows} строк\n"
            f"Брендов: {', '.join(brands_found)}\n"
            f"Генерирую отчёты...",
        )

        # Генерируем и отправляем по каждому бренду
        success_count = 0
        errors = []

        for brand, rows in parsed["by_brand"].items():
            # Пропускаем если нет данных
            if not rows:
                continue

            # Ищем группу для этого бренда
            target_group = BRAND_GROUPS.get(brand) or BRAND_GROUPS.get("Другие")
            if not target_group:
                log.warning(f"  Нет группы для бренда {brand}, пропускаем")
                continue

            log.info(f"  Генерирую отчёт для {brand} ({len(rows)} строк)...")

            try:
                # Генерируем HTML через Claude
                html = generate_brand_html(brand, rows, parsed.get("totals"))

                # Отправляем HTML файлом
                file_name = f"ASTER_{brand}_{datetime.now().strftime('%Y%m%d_%H%M')}.html"
                html_bytes = html.encode("utf-8")

                caption = (
                    f"📊 *{brand}* · Отчёт ADM-ASTER\n"
                    f"📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
                    f"📋 Филиалов: {len(rows)}\n"
                    f"_Откройте файл в браузере_"
                )

                await context.bot.send_document(
                    chat_id=target_group,
                    document=io.BytesIO(html_bytes),
                    filename=file_name,
                    caption=caption,
                    parse_mode=ParseMode.MARKDOWN,
                )

                log.info(f"  ✅ {brand} → группа {target_group}")
                success_count += 1

            except Exception as e:
                err_msg = f"❌ {brand}: {str(e)[:100]}"
                log.error(err_msg)
                errors.append(err_msg)

        # Финальный отчёт администратору
        summary = (
            f"✅ Готово!\n"
            f"Отправлено: {success_count}/{len(parsed['by_brand'])} брендов\n"
            f"Файл: `{filename}`\n"
            f"Время: {datetime.now().strftime('%H:%M:%S')}"
        )
        if errors:
            summary += "\n\nОшибки:\n" + "\n".join(errors)

        await context.bot.send_message(
            ADMIN_CHAT_ID, summary, parse_mode=ParseMode.MARKDOWN
        )

    except Exception as e:
        err = f"❌ Критическая ошибка: {str(e)}"
        log.error(err, exc_info=True)
        await context.bot.send_message(ADMIN_CHAT_ID, err)


# ─── Команды бота ─────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 ADM-ASTER Report Bot\n\n"
        "Жду Excel файл из группы-источника.\n"
        "Как только файл придёт — автоматически разошлю отчёты по брендам.\n\n"
        f"Слежу за группой: {SOURCE_GROUP_ID}"
    )

async def cmd_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    brands_list = "\n".join(f"  • {b}: группа {g}" for b, g in BRAND_GROUPS.items())
    await update.message.reply_text(
        f"✅ Бот работает\n"
        f"Источник: {SOURCE_GROUP_ID}\n\n"
        f"Группы брендов:\n{brands_list}"
    )


# ─── MAIN ─────────────────────────────────────────────────────

def main():
    from telegram.ext import CommandHandler

    print("=" * 55)
    print("  ADM-ASTER Report Bot")
    print(f"  {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
    print("=" * 55)

    if "ВАШ_ТОКЕН" in BOT_TOKEN:
        print("❌ Заполните BOT_TOKEN! Получить у @BotFather")
        return
    if "sk-ant-..." in ANTHROPIC_KEY:
        print("❌ Заполните ANTHROPIC_KEY!")
        return

    app = Application.builder().token(BOT_TOKEN).build()

    # Handlers
    app.add_handler(CommandHandler("start",  cmd_start))
    app.add_handler(CommandHandler("status", cmd_status))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))

    print("✅ Бот запущен. Ожидаю Excel файлы...")
    print("   Нажмите Ctrl+C для остановки\n")

    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
